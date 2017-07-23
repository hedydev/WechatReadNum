
import requests
import json
import re
import openpyxl
import time

def postUrl():
	url = 'https://mp.weixin.qq.com/mp/getappmsgext?__biz=MjM5MjA4MjA4MA==&appmsg_type=9&mid=2654678662&sn=f4af6edec13b870fac7c147884aacc4c&idx=1&scene=0&title=%E2%80%9C%E5%8F%B0%E7%8B%AC%E2%80%9D%E5%AA%92%E4%BD%93%E5%A6%84%E7%A7%B0%E6%94%AF%E6%8C%81%E4%B8%AD%E5%8D%B0%E5%BC%80%E6%88%98%EF%BC%8C%E5%8F%B0%E5%A5%B3%E6%98%9F%E5%8F%91%E6%96%87%E6%80%92%E6%96%A5&ct=1500710188&abtest_cookie=&devicetype=Windows%208&version=/mmbizwap/zh_CN/htmledition/js/appmsg/index373dc8.js&f=json&r=0.9470569686964154&is_need_ad=0&comment_id=3423490704&is_need_reward=0&both_ad=1&reward_uin_count=0&msg_daily_idx=2&is_original=0&uin=MTczMzE5NjY2MQ%253D%253D&key=dd608d91fd702d6455443f0c158586d26272e1248a03182b3a46508d4df8a09b128fdcaee9d4f1a26efe130f68bc873329aa7bb1463104237f50914a47c46fb7e8d438a65a14911b981feb1dd6d976dd&pass_ticket=GktAVKLd2aiz2O5ne6mhTYbwlsm0DR1pWAp%25252FlPqoGytTaYb1QH2soJlkcTFC3pFP&wxtoken=1248054100&devicetype=Windows%26nbsp%3B8&clientversion=62040549&x5=0&f=json'
	
	h = {'Host': 'mp.weixin.qq.com',
			'Connection': 'keep-alive',
			'Accept': '*/*',
			'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
			'Origin': 'https://mp.weixin.qq.com',
			'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36 MicroMessenger/6.5.2.501 NetType/WIFI WindowsWechat QBCore/3.43.556.400 QQBrowser/9.0.2524.400',
			'X-Requested-With': 'XMLHttpRequest',
			'Referer': 'https://mp.weixin.qq.com/s?__biz=MjM5MjA4MjA4MA==&mid=2654678662&idx=1&sn=f4af6edec13b870fac7c147884aacc4c&chksm=bd643f498a13b65f4996de8a5ee82823045469903b8e25772a46bb085e535bd7345da9a3e957&scene=0&key=dd608d91fd702d6455443f0c158586d26272e1248a03182b3a46508d4df8a09b128fdcaee9d4f1a26efe130f68bc873329aa7bb1463104237f50914a47c46fb7e8d438a65a14911b981feb1dd6d976dd&ascene=1&uin=MTczMzE5NjY2MQ%3D%3D&devicetype=Windows+8&version=62040549&pass_ticket=GktAVKLd2aiz2O5ne6mhTYbwlsm0DR1pWAp%2FlPqoGytTaYb1QH2soJlkcTFC3pFP&winzoom=1',
			'Accept-Language': 'zh-CN,zh;q=0.8,en-us;q=0.6,en;q=0.5;q=0.4',
			'Cookie': 'wxtokenkey=8c8d66896cbc9d3a7e9b2241b5693379bf85457e24677e1aeea86a19059e8d42; wxuin=1733196661; pass_ticket=GktAVKLd2aiz2O5ne6mhTYbwlsm0DR1pWAp/lPqoGytTaYb1QH2soJlkcTFC3pFP; wap_sid2=CPX2uboGEnBKU0tDN2hvSV9RUzRadTdCNnQ1THVrelE5a1dmaEU1bG5SYVJNR3pXbExZeDZHMnFFcm5YSVNhZ1E3WWVFLUtYQjU2OE8wbW01QmZiUUpRQ1ZYX0lNU3gwc3FsRTNVcmQ4YkxUVXpYbk5LdVNBd0FBMPubzMsFOA1AAQ=='
	}

	d = {'is_only_read': '1',
		'req_id': 'eq_id=2216goWwOPMRXthjobCvBRH8',
		'pass_ticket': 'GktAVKLd2aiz2O5ne6mhTYbwlsm0DR1pWAp%25252FlPqoGytTaYb1QH2soJlkcTFC3pFP',
		'is_temp_url': '0'
	}

	re = requests.post(url, headers=h, data=d)
	return re.text


def getRedNum(re_text):
	json_data = json.loads(re_text)
	appmsgstat = json_data.get('appmsgstat', -1)
	read_num = appmsgstat.get('read_num', -1)
	print(read_num)
	return read_num

def getLikeNum(re_text):
	json_data = json.loads(re_text)
	appmsgstat = json_data.get('appmsgstat', -1)
	like_num = appmsgstat.get('like_num', -1)
	print(like_num)
	return like_num
	
def sleeptime(hour, min, sec):
    return hour*3600 + min*60 + sec


if __name__ == "__main__":

	wb = openpyxl.Workbook();	#创建工作簿
	sheet = wb.create_sheet('1')
	sheet.cell(row=1, column=1).value = '时间'
	sheet.cell(row=1, column=2).value = '阅读量'
	sheet.cell(row=1, column=3).value = '点赞数'

	currentTime = time.strftime('%Y-%m-%d%H%M',time.localtime(time.time()))

	wbName = 'C:/Users/Administrator/Desktop/观察者网_“台独”媒体妄称支持中印开战，台女星发文怒斥_' + currentTime + '.xlsx'
	wb.save(wbName);

	rows = 2;	#起始行
	pauseTime = sleeptime(0,1,0);	#程序暂停时间

	for i in range(300):
		print(time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time())))

		re = postUrl()
		read = getRedNum(re)
		like = getLikeNum(re)

		wb['1'].cell(row=rows, column=1).value = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()))
		wb['1'].cell(row=rows, column=2).value = read
		wb['1'].cell(row=rows, column=3).value = like

		rows += 1;
		wb.save(wbName);
		time.sleep(pauseTime);