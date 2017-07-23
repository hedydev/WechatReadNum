
import requests
import json
import re
import openpyxl
import time

def postUrl():
	url = 'https://mp.weixin.qq.com/mp/getappmsgext?__biz=MjM5MzI5NTU3MQ==&appmsg_type=9&mid=2651457113&sn=b040cdeaf045fdb0db642fa87f984343&idx=1&scene=0&title=%E2%80%9C%E5%BC%BA%E5%A5%B8%E5%BC%8F%E5%AE%89%E6%85%B0%E2%80%9D%EF%BC%9F%EF%BC%81%E6%B9%96%E5%8D%97%E4%B8%80%E5%9C%A8%E8%AF%BB%E5%8D%9A%E5%A3%AB%E6%B6%89%E6%A1%88%E8%A2%AB%E8%B0%83%E6%9F%A5&ct=1500731984&abtest_cookie=&devicetype=Windows%208&version=/mmbizwap/zh_CN/htmledition/js/appmsg/index373dc8.js&f=json&r=0.7715636789798737&is_need_ad=0&comment_id=172783341&is_need_reward=0&both_ad=1&reward_uin_count=0&msg_daily_idx=3&is_original=0&uin=MTczMzE5NjY2MQ%253D%253D&key=41aba78a08e0c639c51289f9161db8f0e9440c42a0cbf92776a93eca628c00a4c69aa564b766aefbd97686ac42ebf4d444332f0def3afff82a302bd82bda6624f0d970a41fa683345491002a625d73c2&pass_ticket=GktAVKLd2aiz2O5ne6mhTYbwlsm0DR1pWAp%25252FlPqoGytTaYb1QH2soJlkcTFC3pFP&wxtoken=2028425923&devicetype=Windows%26nbsp%3B8&clientversion=62040549&x5=0&f=json'
	h = {'Host': 'mp.weixin.qq.com',
			'Connection': 'keep-alive',
			'Accept': '*/*',
			'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
			'Origin': 'https://mp.weixin.qq.com',
			'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36 MicroMessenger/6.5.2.501 NetType/WIFI WindowsWechat QBCore/3.43.556.400 QQBrowser/9.0.2524.400',
			'X-Requested-With': 'XMLHttpRequest',
			'Referer': 'https://mp.weixin.qq.com/s?__biz=MjM5MzI5NTU3MQ==&mid=2651457113&idx=1&sn=b040cdeaf045fdb0db642fa87f984343&chksm=bd679be58a1012f3780f78178e160c44e356168e00c61e7c72a6c749982111c7cc27a7f10ac3&scene=0&key=41aba78a08e0c639c51289f9161db8f0e9440c42a0cbf92776a93eca628c00a4c69aa564b766aefbd97686ac42ebf4d444332f0def3afff82a302bd82bda6624f0d970a41fa683345491002a625d73c2&ascene=1&uin=MTczMzE5NjY2MQ%3D%3D&devicetype=Windows+8&version=62040549&pass_ticket=GktAVKLd2aiz2O5ne6mhTYbwlsm0DR1pWAp%2FlPqoGytTaYb1QH2soJlkcTFC3pFP&winzoom=1',
			'Accept-Language': 'zh-CN,zh;q=0.8,en-us;q=0.6,en;q=0.5;q=0.4',
			'Cookie': 'wxtokenkey=51c1b109485200c94a5ee22c766ffa36f5efa5d33af20f55837eba95c015df39; wxuin=1733196661; pass_ticket=GktAVKLd2aiz2O5ne6mhTYbwlsm0DR1pWAp/lPqoGytTaYb1QH2soJlkcTFC3pFP; wap_sid2=CPX2uboGEnBKU0tDN2hvSV9RUzRadTdCNnQ1THVoUDE1d2dRa096NlVHWWpzSEFERlRSMENfNHF1OW55MU1vRGhDdk5RaVY2ZUt6SXI1YnlLeE40OVVPS1VYME1lRS13azRuSEdWeWJrNjhSYVU4OEpibVNBd0FBMPS2zcsFOA1AAQ=='
	}

	d = {'is_only_read': '1',
		'req_id': 'eq_id=2222oxziZaFmynF1HNJGU15s',
		'pass_ticket': 'GktAVKLd2aiz2O5ne6mhTYbwlsm0DR1pWAp%25252FlPqoGytTaYb1QH2soJlkcTFC3pFP',
		'is_temp_url': '0'
	}

	re = requests.post(url, headers=h, data=d)
	return re.text


def getRedNum(re_text):
	json_data = json.loads(re_text)
	appmsgstat = json_data.get('appmsgstat', -1)
	read_num = appmsgstat.get('read_num', -1)
	print(read_num)
	return read_num

def getLikeNum(re_text):
	json_data = json.loads(re_text)
	appmsgstat = json_data.get('appmsgstat', -1)
	like_num = appmsgstat.get('like_num', -1)
	print(like_num)
	return like_num
	
def sleeptime(hour, min, sec):
    return hour*3600 + min*60 + sec


if __name__ == "__main__":

	wb = openpyxl.Workbook();	#创建工作簿
	sheet = wb.create_sheet('1')
	sheet.cell(row=1, column=1).value = '时间'
	sheet.cell(row=1, column=2).value = '阅读量'
	sheet.cell(row=1, column=3).value = '点赞数'

	currentTime = time.strftime('%Y-%m-%d%H%M',time.localtime(time.time()))

	wbName = 'C:/Users/Administrator/Desktop/澎湃新闻_强奸式安慰' + currentTime + '.xlsx'
	wb.save(wbName);

	rows = 2;	#起始行
	pauseTime = sleeptime(0,1,0);	#程序暂停时间

	for i in range(300):
		print(time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time())))

		re = postUrl()
		read = getRedNum(re)
		like = getLikeNum(re)

		wb['1'].cell(row=rows, column=1).value = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()))
		wb['1'].cell(row=rows, column=2).value = read
		wb['1'].cell(row=rows, column=3).value = like

		rows += 1;
		wb.save(wbName);
		time.sleep(pauseTime);
	
	